import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
import numpy as np

wb = load_workbook('2020_kosdaq.xlsx')
ws_names = wb.sheetnames
ws_names = ws_names[1:]

pred_list = []

for name in ws_names:
    ws = wb[name]
    #print(ws['G24'].value)
    pred_list.append(ws['G24'].value)
 
wb = load_workbook('2021_kosdaq.xlsx')
ws_names = wb.sheetnames
ws_names = ws_names[1:]

curr_list = []

for name in ws_names:
    ws = wb[name]
    #print(ws['G15'].value)
    curr_list.append(ws['G15'].value)


res = []
res_mod = []
for pred,curr in zip(pred_list, curr_list):
    res.append((curr-pred)/curr)
    res_mod.append((curr-(pred*0.7))/curr)
    #print(pred, curr)

print(np.mean(res))
print(np.mean(res_mod))


res = []

w = 0.7
for name in ws_names:
    ws = wb[name]
    list_mean = []
    
    curr = ws['D15'].value
    pred = ws['D24'].value
    list_mean.append((curr-(pred*w))/curr)
    
    curr = ws['E15'].value
    pred = ws['E24'].value
    list_mean.append((curr-(pred*w))/curr)

    curr = ws['F15'].value
    pred = ws['F24'].value
    list_mean.append((curr-(pred*w))/curr)

    curr = ws['G15'].value
    pred = ws['G24'].value
    list_mean.append((curr-(pred*w))/curr)

    res.append(np.mean(list_mean))


print(np.mean(res))









    

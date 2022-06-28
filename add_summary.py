import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

#data = pd.read_excel('2022-06-20.xlsx', engine='openpyxl', sheet_name=None)
#print(data)

#for d in data:
#    print(data[d]['i16'])


#sheet_names = []
#for d in data:
#    sheet_names.append(d)

#print(sheet_names)

#print(data['셀트리온']['Pred'][14])

wb = load_workbook('2022-06-24.xlsx')

ws_names = wb.sheetnames

ws_names = ws_names[1:]
#print(ws_names)

ws = wb[ws_names[1]]
print(ws['I17'].value)

write_wb = Workbook()
write_ws = write_wb.create_sheet('Summary')

write_ws = write_wb.active
write_ws.append(['','매출증가율','영업이익증가율','영업현금증가율','평균증가율','주가증가율','시총증가율','평균증가율','예측가격','매수가격','상태'])

base_g = 20

for name in ws_names:
    info = []
    ws = wb[name]
    info.append(name)
    info.append(ws['I17'].value)
    info.append(ws['I18'].value)
    info.append(ws['I19'].value)
    info.append((ws['I17'].value+ws['I18'].value)/2)
    info.append(ws['I20'].value)
    info.append(ws['I21'].value)
    info.append((ws['I20'].value+ws['I21'].value)/2)
    info.append(ws['I24'].value)
    if ws['I24'].value is None:
        info.append(0)
    else:
        info.append(ws['I24'].value*0.7)

    if ws['i17'].value > base_g and ws['i18'].value > base_g and ws['i19'].value > base_g:
        if ((ws['I17'].value+ws['I18'].value)/2) > ((ws['I20'].value+ws['I21'].value)/2):
            info.append('매수')
        else:
            info.append('대기')
    else:
        info.append('보류')
       

    write_ws.append(info)
write_wb.save('summary.xlsx')

